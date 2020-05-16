import openpyxl
import xlsxwriter 
import re
from pathlib import Path

xlsx_file = Path('Sheets Election Week 5-1 (Responses).xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file) 
workbook = xlsxwriter.Workbook('results.xlsx')
worksheet = workbook.add_worksheet()
quickie = ["Etheldrea", "Deymos", "Exodo", "Amaterasu", "Dorjan", "Gaul", "Kyubi", "Gorey", "Hiro", "Karla", "Marchosias", "Mazur", "Helena", "Tawil", "Valis", "Alystaire", "Angelus", "Dereck", "Evermillion"]

# Read the active sheet:
sheet = wb_obj.active
names = set()
results = dict()

for row in sheet.iter_rows(min_row = 2, max_row=250, max_col=5, min_col=3):
    vote = set()
    for cell in row:
        if cell.value != None:
            v = cell.value.lower().strip()
            va = re.sub(r'[àá]',r'a', v)
            ve = re.sub(r'[èé]',r'e', va)
            vi = re.sub(r'[ìí]',r'i', ve)
            vo = re.sub(r'[òó]',r'o', vi)
            vu = re.sub(r'[ùú]',r'e', vo)
            def replace(match):
                return match.group(1)+match.group(2).upper()
            capitalized = re.sub(r'([ (])+([a-z])', replace, vu)
            r = capitalized[0].upper() + capitalized[1:]
            vote.add(r)
    
    for v in vote:
        for q in quickie:
            if q in v:
                v = q
        names.add(v)
        if v in results:
            results[v] += 1
        else:
            results[v] = 1

row = 0
for i in sorted(results):
    worksheet.write(row, 0, i)
    worksheet.write(row, 1, results[i])
    row += 1

workbook.close()


    
